"""
This module provides classes to handle the writing and manipulation of Excel files
using the openpyxl and pandas libraries.

Classes:
    - ChartLabels: A dataclass to define labels for charts (title, x-axis, y-axis).
    - XlSheetWriter: A class to handle writing and formatting Excel sheets.
    - XlChartWriter: A subclass of XlSheetWriter to handle writing Excel sheets with charts.
    - XlWriter: A class to manage multiple sheets and save them to an Excel file.

Usage:
    The `XlWriter` class is used to write data to multiple sheets, with options for formatting
    and adding charts. The `XlSheetWriter` handles sheet-specific operations, while `XlChartWriter`
    provides chart functionality.

Example:
    xl_writer = XlWriter('path_to_excel_file.xlsx')
    sheet = xl_writer.add_sheet('Sheet1', data_frame)
    xl_writer.save()

Dependencies:
    - openpyxl: For writing and managing Excel files.
    - pandas: For data handling.
    - shared.log: For logging operations.
"""

from dataclasses import dataclass

import openpyxl
import pandas as pd
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from shared import log


@dataclass
class ChartLabels:
    """
    A dataclass to hold chart labels.

    Attributes:
        title (str): The title of the chart.
        x_label (str): The label for the x-axis.
        y_label (str): The label for the y-axis.
    """

    title: str
    x_label: str
    y_label: str


class XlSheetWriter:
    """
    Base class to handle Excel sheet operations, including writing data,
    formatting the sheet, and adjusting column widths.

    Args:
        writer (pd.ExcelWriter): The Excel writer object to write the data.
        sheet_name (str): The name of the sheet.
        df (pd.DataFrame): The DataFrame to be written into the sheet. Defaults to an
        empty DataFrame.
    """

    def __init__(self, writer, sheet_name="Sheet1", df=None):
        if df is None:
            df = pd.DataFrame()
        self.writer = writer
        self.sheet_name = sheet_name
        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False)  # Creates the sheet
        if isinstance(writer.book, openpyxl.Workbook):
            self.ws = writer.book[sheet_name]
        else:
            raise NotImplementedError(
                "This operation is only supported with openpyxl.")

    def adjust_column_width(self, max_number_width=8):
        """
        Adjust the column widths based on the content of the cells.

        Args:
            max_number_width (int): The minimum width for the columns. Default is 8.
        """
        for i, column_cells in enumerate(self.ws.columns, 1):
            max_length = max((len(str(cell.value))
                             for cell in column_cells), default=0)
            self.ws.column_dimensions[get_column_letter(i)].width = max(
                max_length, max_number_width
            )

    def format_worksheet(self):
        """
        Apply basic formatting to the worksheet, including borders, font styles, and alignment.
        """
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in self.ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="FFE7E6E6", end_color="FFE7E6E6", fill_type="solid"
        )
        medium_border = Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium"),
        )

        for i in range(1, self.ws.max_row + 1):
            self.ws.cell(row=i, column=1).border = medium_border
            self.ws.cell(row=i, column=1).font = header_font
            self.ws.cell(row=i, column=1).fill = header_fill
            self.ws.cell(
                row=i, column=1).alignment = Alignment(
                horizontal="left")

        for j in range(1, self.ws.max_column + 1):
            self.ws.cell(row=1, column=j).border = medium_border
            self.ws.cell(row=1, column=j).font = header_font
            self.ws.cell(row=1, column=j).fill = header_fill

        for j in range(2, self.ws.max_column + 1):
            self.ws.cell(
                row=1, column=j).alignment = Alignment(
                horizontal="center")

        self.ws.cell(
            row=1, column=1).alignment = Alignment(
            horizontal="center")

    def finalize_sheet(
            self,
            title="",
            portrait=True,
            fit_to_width=1,
            fit_to_height=0):
        """
        Finalize the sheet by applying formatting, adjusting print settings, and defining
        headers/footers.

        Args:
            title (str): The title to set in the header.
            portrait (bool): Whether to set the sheet to portrait mode.
            fit_to_width (int): Fit the sheet to the specified number of pages in width.
            fit_to_height (int): Fit the sheet to the specified number of pages in height.
        """
        self.format_worksheet()
        self.page_print_setting(
            portrait=portrait,
            fit_to_width=fit_to_width,
            fit_to_height=fit_to_height)
        self.adjust_column_width()
        if title == "":
            title = self.sheet_name
        self.define_header_and_footer(title)

    def page_print_setting(
            self,
            portrait=True,
            fit_to_width=1,
            fit_to_height=0):
        """
        Configure the print settings for the worksheet, including page orientation and
        fit-to-page settings.

        Args:
            portrait (bool): Whether to set the page orientation to portrait.
            fit_to_width (int): Fit the sheet to the specified number of pages in width.
            fit_to_height (int): Fit the sheet to the specified number of pages in height.
        """
        self.ws.page_setup.paperSize = self.ws.PAPERSIZE_A4
        self.ws.page_setup.orientation = (
            self.ws.ORIENTATION_PORTRAIT if portrait else self.ws.ORIENTATION_LANDSCAPE)
        self.ws.page_setup.fitToPage = True
        self.ws.page_setup.fitToWidth = fit_to_width
        self.ws.page_setup.fitToHeight = fit_to_height
        self.ws.page_margins.left = 1
        self.ws.page_margins.right = 1
        self.ws.page_margins.top = 1.5
        self.ws.page_margins.bottom = 1.0
        self.ws.page_margins.header = 0.3
        self.ws.page_margins.footer = 0.3

    def define_header_and_footer(self, title):
        """
        Set the header and footer for the worksheet.

        Args:
            title (str): The title to display in the header.
        """
        footer_text = '&L&"Arial"&12&F &C&"Arial"&12&A  &R&"Arial"&12&P/&N'
        header_text = f'&C&"Arial,Bold"&16{title}&R&"Arial"&12&D'

        self.ws.oddHeader.center.text = header_text
        self.ws.oddFooter.center.text = footer_text
        self.ws.evenHeader.center.text = header_text
        self.ws.evenFooter.center.text = footer_text


class XlChartWriter(XlSheetWriter):
    """
    A class to handle writing a chart to an Excel sheet, extending the XlSheetWriter functionality.

    Args:
        writer (pd.ExcelWriter): The Excel writer object.
        data_work_sheet (XlSheetWriter): The sheet containing the data for the chart.
        chart_sheet_name (str): The name of the chart sheet.
        labels (ChartLabels): An object containing chart labels (title, x-axis, y-axis).
    """

    def __init__(self, writer, data_work_sheet, chart_sheet_name, labels):
        """
        Initializes the ChartCreator class with the writer, data worksheet, chart sheet name, and
        labels.

        Args:
            writer: The Excel writer object for writing the chart.
            data_work_sheet: The worksheet object containing the data to be plotted.
            chart_sheet_name (str): The name of the chart sheet.
            labels (ChartLabels): An object containing the title and labels for the chart.
        """
        self.data_sheet = data_work_sheet
        self.labels = labels
        self.chart = None
        super().__init__(writer, chart_sheet_name)

    def init_global_params(self):
        """
        Initializes the global parameters for the chart, including title, style, and axis labels.
        Sets the categories for the chart based on the data in the worksheet.
        """
        self.chart = LineChart()
        self.chart.title = self.labels.title
        self.chart.style = 10
        self.chart.x_axis.title = self.labels.x_label
        self.chart.y_axis.title = self.labels.y_label

        categories = Reference(
            self.data_sheet.ws,
            min_col=2,
            min_row=1,
            max_row=1,
            max_col=self.data_sheet.ws.max_column,
        )
        self.chart.set_categories(categories)

    def finalize_chart(self):
        """
        Finalizes the chart by formatting it, adding it to the worksheet, and applying
        print settings.
        Also defines the header and footer based on the chart title.
        """
        self.format_chart()
        self.ws.add_chart(self.chart, "A1")
        self.page_print_setting(portrait=False)
        self.define_header_and_footer(self.labels.title)

    def add_chart_data(self, hidden_rows=0):
        """
        Adds data to the chart by iterating over the rows of the data worksheet.
        It takes into account hidden rows, if any, and appends the data series to the chart.

        Args:
            hidden_rows (int, optional): The number of hidden rows to exclude from the chart.
            Defaults to 0.
        """
        last_row = self.data_sheet.ws.max_row - hidden_rows
        # Créez une référence pour les années (colonnes de la première ligne après le nom)
        x_labels = Reference(
            self.data_sheet.ws,
            min_col=2,
            max_col=self.data_sheet.ws.max_column,
            min_row=1
        )

        for row in range(2, last_row + 1):
            values = Reference(
                self.data_sheet.ws,
                min_col=2,
                min_row=row,
                max_row=row,
                max_col=self.data_sheet.ws.max_column,
            )
            series = Series(values, title=self.data_sheet.ws.cell(row, 1).value)
            self.chart.series.append(series)
            series.smooth = False

        # Associez les années comme étiquettes pour l'axe X
        self.chart.set_categories(x_labels)

    def create_chart(self):
        """
        Creates a line chart based on the data in the worksheet. It initializes the chart,
        adds data series to it, and finalizes the chart by applying all settings.
        """
        self.init_global_params()

        # method to overload if necessary
        self.add_chart_data()
        self.finalize_chart()

    def format_chart(self):
        """
        Apply formatting to the chart, including layout and axis settings.
        """
        self.chart.width = 1300 / 72 * 2.54
        self.chart.height = 800 / 72 * 2.54

        self.chart.y_axis.majorGridlines = ChartLines()
        self.chart.y_axis.minorGridlines = ChartLines()
        self.chart.y_axis.minorTickMark = "out"

        self.chart.x_axis.minorGridlines = ChartLines()
        self.chart.x_axis.minorTickMark = "out"
        self.chart.x_axis.tickLblPos = "low"
        self.chart.y_axis.tickLblPos = "low"

        self.chart.x_axis.number_format = "General"
        self.chart.y_axis.number_format = "General"

        self.chart.x_axis.tickLblSkip = 1
        self.chart.y_axis.tickLblSkip = 1

        self.chart.layout = Layout(
            manualLayout=ManualLayout(x=0.005, y=0.005, h=0.90, w=0.85)
        )


class XlWriter:
    """
    A class to manage writing multiple sheets (including charts) to an Excel file.

    Args:
        xl_file (str): The path to the Excel file.
    """

    # pylint: disable=abstract-class-instantiated
    def __init__(self, xl_file):
        self.sheets = []
        self.xl_file = xl_file
        self.chart_writer = XlChartWriter
        self.writer = pd.ExcelWriter(xl_file, engine="openpyxl")

    def add_sheet(self, sheet_name, df=None):
        """
        Add a new sheet to the Excel file.

        Args:
            sheet_name (str): The name of the new sheet.
            df (pd.DataFrame): The data to write to the sheet. Defaults to None.

        Returns:
            XlSheetWriter: The sheet writer object if the sheet is created; None otherwise.
        """
        if self.get_sheet(sheet_name) is None:
            sheet = XlSheetWriter(self.writer, sheet_name, df)
            self.sheets.append(sheet)
            return sheet
        return None

    def add_chart_sheet(self, data_worksheet, chart_sheet_name, labels):
        """
        Add a new sheet with a chart to the Excel file.

        Args:
            data_worksheet (XlSheetWriter): The worksheet containing the data for the chart.
            chart_sheet_name (str): The name of the chart sheet.
            labels (ChartLabels): The labels for the chart.

        Returns:
            XlChartWriter: The chart writer object if the sheet is created; None otherwise.
        """
        if self.get_sheet(chart_sheet_name) is None:
            sheet = self.chart_writer(
                self.writer, data_worksheet, chart_sheet_name, labels
            )
            self.sheets.append(sheet)
            return sheet
        return None

    def get_sheet(self, sheet_name):
        """
        Retrieve an existing sheet by name.

        Args:
            sheet_name (str): The name of the sheet to retrieve.

        Returns:
            XlSheetWriter: The sheet writer object if the sheet is found; None otherwise.
        """
        for sheet in self.sheets:
            if sheet.ws.title == sheet_name:
                return sheet
        return None

    def save(self):
        """
        Save the Excel file to the specified path.
        """
        try:
            if not self.writer.book.sheetnames:
                raise ValueError("No sheets found in the workbook.")
            self.writer.close()
            log.info("%s successfully created.", self.xl_file)
        except ValueError as e:
            log.error("Error: %s", e)
        except (IOError, OSError) as e:
            log.error("File error while saving: %s", e)
        # pylint: disable = broad-exception-caught
        except Exception as e:
            log.error("An unexpected error occurred: %s", e)
