"""
This module provides classes and functions to create and manipulate pivot tables
and charts in Excel files using the `openpyxl` and `pandas` libraries. It
includes functionality to generate pivot tables, evaluate formulas, and export
charts, customized for specific criteria.

Classes:
    XlPivotChartWriter: Handles the creation of pivot charts.
    XlPivotWriter: Manages the creation of pivot tables and exports charts.
"""

import re

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.chart import Reference, Series
from shared import project
from xl.xl_writer import ChartLabels, XlChartWriter, XlSheetWriter, XlWriter


class XlPivotChartWriter(XlChartWriter):
    """
    Customizes the behavior for creating pivot charts in Excel.

    Args:
        XlChartWriter: Inherits base chart functionalities.
    """

    def __init__(self, writer, data_work_sheet, chart_sheet_name, labels):
        self.show_rows = None
        self.show_total = None
        self.show_delta = None
        self.show_init = None
        super().__init__(writer, data_work_sheet, chart_sheet_name, labels)

    def add_chart_data(self, hidden_rows=0):
        hidden_rows = int(self.show_total) + \
            int(self.show_delta) + int(self.show_init)
        last_row = self.data_sheet.ws.max_row
        if self.show_rows:
            super().add_chart_data(hidden_rows)

        if self.show_total:
            total_row = last_row + 1 - hidden_rows
            values = Reference(
                self.data_sheet.ws,
                min_col=2,
                min_row=total_row,
                max_row=total_row,
                max_col=self.data_sheet.ws.max_column,
            )
            series = Series(values, title="Total")
            self.chart.series.append(series)
            series.smooth = False

        if self.show_delta:
            delta_row = last_row + 1 - hidden_rows + 1
            values = Reference(
                self.data_sheet.ws,
                min_col=2,
                min_row=delta_row,
                max_row=delta_row,
                max_col=self.data_sheet.ws.max_column,
            )
            series = Series(values, title="Delta")
            self.chart.series.append(series)
            series.smooth = False


class XlPivotWriter(XlWriter):
    """
    Class to handle operations related to pivot tables and exporting charts in Excel files.

    Args:
        xl_file (str): The path to the Excel file.
    """

    def __init__(self, xl_file):
        super().__init__(xl_file)
        self.chart_writer = XlPivotChartWriter

    def create_criterion_pivots(self, data_df):
        """
        Creates pivot tables for each unique criterion key in the provided DataFrame.

        Args:
            data_df (DataFrame): The DataFrame containing the data to pivot.
        Returns:
            tuple: A dictionary of pivot tables indexed by criterion, and a list of criteria.
        """

        criterion_pivots = {}
        criteria = data_df["criterion_key"].unique()
        all_columns = data_df["columns"].unique().tolist()

        for criterion in criteria:
            filtered_df = data_df[data_df["criterion_key"] == criterion]
            pivot_table = filtered_df.pivot(
                index="index",
                columns="columns",
                values="value",
            )
            pivot_table = pivot_table.reindex(
                columns=all_columns, fill_value=0)
            criterion_pivots[criterion] = pivot_table
        return criterion_pivots, criteria

    def create_pivot_tables(
        self,
        data_df,
        pivot_information_df,
    ):
        """
        Creates pivot tables for each formula in the pivot information DataFrame.

        Args:
            data_df (DataFrame): The DataFrame containing the raw data.
            pivot_information_df (DataFrame): DataFrame containing formulas and
            information about each pivot.
        """
        criterion_pivots, criteria = self.create_criterion_pivots(data_df)
        formulas_df = pivot_information_df

        for index, row in formulas_df.iterrows():
            _ = index
            formula = row["formula"]
            if pd.isna(formula):
                continue

            query_name = row["query_name"]
            language = project.context.language
            data_sheet_name = (
                project.get_resource_string(
                    f"{query_name}_Sheet_Prefix",
                    language) + "_Data")
            result_df = self.process_formula(
                criterion_pivots, criteria, formula)
            # suppress empty columns
            result_df = result_df.loc[:, ~(
                (result_df.iloc[1:].replace(0, float('NaN')).isnull()).all())]
            result_df = result_df.reset_index()
            sh = XlSheetWriter(self.writer, data_sheet_name, result_df)

            self.finalize_data_sheet(sh, row)
            self.export_chart(sh, row)

    def sort_key(self, x):
        """
        Sorting key function for custom sorting of pivot table columns.

        Args:
            x: The value to sort.

        Returns:
            int or str: The sorting key.
        """
        try:
            return int(x)
        except ValueError:
            try:
                return int(x[2:])
            except (ValueError, IndexError):
                return x

    def process_formula(self, pivot_tables, criteria, formula):
        """
        Processes a given formula and evaluates it against the provided pivot tables.

        Args:
            pivot_tables (dict): Dictionary of pivot tables.
            criteria (list): List of criteria to use for processing.
            formula (str): The formula to evaluate.

        Returns:
            DataFrame: A DataFrame containing the evaluated result.
        """
        result_data = []
        for index in pivot_tables[criteria[0]].index:
            result_row = {"name": index}
            for column in pivot_tables[criteria[0]].columns:
                result = self.eval_formula(
                    pivot_tables, formula, index, column)
                if not pd.isna(result):
                    result_row[column] = result
            result_data.append(result_row)

        result_df = pd.DataFrame(result_data).set_index("name")
        result_df = result_df.dropna(axis=1, how="all")
        result_df = result_df.loc[:, (result_df != 0).any(axis=0)]
        result_df = result_df.loc[:, ~result_df.isna().all(axis=0)]
        result_df = result_df.reindex(
            sorted(result_df.columns, key=self.sort_key), axis=1
        )
        return result_df

    def eval_formula(self, pivot_tables, formula, index, column):
        """
        Evaluates a custom formula using the data from pivot tables.

        Args:
            pivot_tables (dict): Dictionary of pivot tables.
            formula (str): The formula string to evaluate.
            index: The row (index) to look up.
            column: The column (column) to look up.

        Returns:
            float: The evaluated result of the formula, or NaN if an error occurs.
        """
        tokens = re.findall(
            r"[\d.]+(?:[eE][+-]?\d+)?|[+\-*/()]|[\w.]+", formula)
        for i, token in enumerate(tokens):
            if re.match(r"^\d", token):
                tokens[i] = token.replace(",", ".")
            elif re.match(r"^[a-zA-Z]", token):
                criterion = token
                if (
                    criterion in pivot_tables
                    and column in pivot_tables[criterion].columns
                    and index in pivot_tables[criterion].index
                ):
                    value = pivot_tables[criterion].at[index, column]
                    if pd.isna(value):
                        return np.nan
                    tokens[i] = str(value) if value is not None else "0"
                else:
                    tokens[i] = "0"

        final_formula = "".join(tokens)
        try:
            return eval(final_formula)  # pylint: disable=eval-used
        except ZeroDivisionError:
            return np.nan
        except NameError as e:
            print(f"Error evaluating formula: {final_formula}, {e}")
            return np.nan

    def add_index_sheet(self, pivot_infos_df):
        """
        Creates an index sheet listing the pivot information.

        Args:
            pivot_infos_df (DataFrame): DataFrame containing information about the pivots.
        """
        sheet_name = "Index"
        columns_to_drop = [
            "id",
            "show_rows",
            "show_total",
            "show_delta",
            "show_init"
        ]
        df = pivot_infos_df.drop(columns=columns_to_drop)

        # add missing columns from the ResourceStrings
        df["sheet_prefix"] = df["query_name"].apply(project.get_sheet_prefix)
        df["title"] = df["query_name"].apply(project.get_sheet_title)

        # reorder the columns
        desired_order = ["title", "sheet_prefix", "formula", "query_name"]
        df = df[desired_order]

        df = df.sort_values(by="title", ascending=True)
        sh = XlSheetWriter(self.writer, sheet_name, df)
        sh.finalize_sheet(portrait=False, title="Index")
        return sh

    def finalize_data_sheet(self, sh, row):
        """
        Finalizes the data sheet by adding total and delta rows if specified.

        Args:
            sh (XlSheetWriter): The sheet writer object.
            row (Series): The row containing the specifications for the sheet.
        """

        if not row["show_init"]:
            self.remove_init_row(sh.ws)

        if row["show_total"]:
            self.add_total_row(sh.ws)

        if row["show_delta"]:
            self.add_delta_row(sh.ws)

        sheet_base_name = row["query_name"]
        sh.finalize_sheet(
            portrait=False,
            title=project.this_db.get_resource_string(
                f"{sheet_base_name}_Title",
                "en"),
        )

    def add_delta_row(self, ws):
        """
        Adds a row to the sheet that calculates the delta (difference) between columns.

        Args:
            ws (Worksheet): The worksheet object.
        """
        last_row = ws.max_row
        last_col = ws.max_column

        ws.cell(row=last_row + 1, column=1, value="Delta")

        for col in range(3, last_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
            delta_formula = f"={col_letter}{last_row}-{prev_col_letter}{last_row}"
            ws.cell(row=last_row + 1, column=col, value=delta_formula)

    def remove_init_row(self, ws):
        """
        Adds a row to the sheet that calculates the total for each column.

        Args:
            ws (Worksheet): The worksheet object.
        """
        last_row = ws.max_row
        if ws.cell(row=last_row, column=1).value == "zz_Init":
            ws.delete_rows(last_row)

    def add_total_row(self, ws):
        """
        Adds a row to the sheet that calculates the total for each column.

        Args:
            ws (Worksheet): The worksheet object.
        """
        last_row = ws.max_row
        last_col = ws.max_column

        ws.cell(row=last_row + 1, column=1, value="Total")

        for col in range(2, last_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            sum_formula = f"=SUM({col_letter}2:{col_letter}{last_row})"
            ws.cell(row=last_row + 1, column=col, value=sum_formula)

    def export_chart(self, data_sheet, row):
        """
        Exports a chart based on the data in the provided sheet.

        Args:
            data_sheet (XlSheetWriter): The sheet containing the data for the chart.
            row (Series): The row containing information about the pivot and the graph.

        Returns:
            XlSheetWriter: The chart sheet object.
        """
        query_name = row["query_name"]
        language = project.context.language

        chart_sheet_name = (
            project.get_resource_string(f"{query_name}_Sheet_Prefix", language)
            + "_Chart"
        )

        labels = ChartLabels(
            title=project.get_resource_string(f"{query_name}_Title", language),
            x_label=project.get_resource_string(f"{query_name}_X_Label", language),
            y_label=project.get_resource_string(f"{query_name}_Y_Label", language),
        )

        sh = self.add_chart_sheet(data_sheet, chart_sheet_name, labels)
        sh.show_rows = row["show_rows"]
        sh.show_total = row["show_total"]
        sh.show_delta = row["show_delta"]
        sh.show_init = row["show_delta"]
        sh.create_chart()
        return sh
