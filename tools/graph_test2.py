# -*- coding: utf-8 -*-
# pylint: disable=abstract-class-instantiated
# pylint causes a false positive
"""
This script generates an Excel file with data and a line chart.

The data is provided as a DataFrame, and the script writes it to an Excel file
and generates a line chart representing the data. The chart is saved on the
same sheet within the workbook.

The output Excel file is saved in the 'data/output' directory relative
to the script's location.
"""

import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.utils.dataframe import dataframe_to_rows

# Get the current script directory
script_dir = os.path.dirname(os.path.abspath(__file__))

# Specify the output file path relative to the script directory
output_file_path = os.path.join(
    script_dir, '..', 'data', 'output', 'test_graph2.xlsx'
)

# Example DataFrame (replace with your actual DataFrame)
data = {
    'Year': ['2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', '2023'],
    'Davos': [1, 2, 3, 4, 5, 6, 7, 8, 9],
    'Baden': [2, 3, 4, 5, 6, 7, 8, 9, 10],
    'Lugano': [3, 4, 5, 6, 7, 8, 9, 10, 11]
}
df = pd.DataFrame(data)

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Write the DataFrame to the worksheet
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Create a LineChart
chart = LineChart()
chart.title = "Casino Revenue Over Years"
chart.style = 10
chart.x_axis.title = "Year"
chart.y_axis.title = "Revenue"

# Define categories (X axis values) using the first column (Year)
categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

# Add the series to the chart
for i in range(2, ws.max_column + 1):
    values = Reference(ws, min_col=i, min_row=2, max_row=ws.max_row)
    series = Series(values, title=ws.cell(row=1, column=i).value)
    chart.series.append(series)

chart.set_categories(categories)

# Set the size of the chart in centimeters
chart.width = 15  # Width in cm
chart.height = 10  # Height in cm

# Ensure axis labels and gridlines are displayed
chart.x_axis.majorGridlines = ChartLines()
chart.y_axis.majorGridlines = ChartLines()
chart.x_axis.tickLblPos = 'nextTo'
chart.y_axis.tickLblPos = 'nextTo'

# Position the chart on the sheet
ws.add_chart(chart, "E5")  # Position the chart starting from cell E5

# Save the Excel file
wb.save(output_file_path)
