# -*- coding: utf-8 -*-
# pylint: disable=abstract-class-instantiated
# pylint causes a false positive
"""
This script generates an Excel file with data and a line chart.

The data is provided as a DataFrame, and the script transposes the data,
writes it to an Excel file, and generates a line chart representing the
data. The chart is saved on a separate sheet within the same workbook.

The output Excel file is saved in the 'data/output' directory relative
to the script's location.
"""

import os

import pandas as pd
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout

# Example DataFrame (replace with your actual DataFrame)
data = {
    'Year': ['2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', '2023'],
    'Bad Ragaz': [1, 2, 3, 4, 5, 6, 7, 8, 9],
    'Baden': [2, 3, 4, 5, 6, 7, 8, 9, 10],
    'Basel': [3, 4, 5, 6, 7, 8, 9, 10, 11]
}

# Convert the dictionary to a DataFrame
df = pd.DataFrame(data)
print(f"\n{df.head()}\n")

# Transpose the DataFrame without adding an index row
df = df.T
df.reset_index(inplace=True)
print(f"\n{df.head()}\n")

# Set the first row as the header
df.columns = df.iloc[0]
print(f"\n{df.head()}\n")

# Remove the first row
df = df[1:]
print(f"\n{df.head()}\n")

# Get the current script directory
script_dir = os.path.dirname(os.path.abspath(__file__))
print(script_dir)

# Specify the output file path relative to the script directory
output_file_path = os.path.join(
    script_dir, '..', 'data', 'output', 'test_graph.xlsx'
)

# Open the Excel file using pandas.ExcelWriter
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    # Write the DataFrame to the Excel file
    SHEET_NAME = 'Data'
    df.to_excel(writer, sheet_name=SHEET_NAME, index=False)

    # Get the workbook and worksheet objects
    wb = writer.book
    ws = wb[SHEET_NAME]

    # Add a new sheet for the chart
    SHEET_PREFIX = 'SheetPrefix'
    chart_sheet_name = f'{SHEET_PREFIX}_Chart'
    i = 1
    tmp = chart_sheet_name
    while tmp in wb.sheetnames:
        tmp = f'{chart_sheet_name}_{i}'
        i += 1
    chart_sheet_name = tmp
    ws_chart = wb.create_sheet(title=chart_sheet_name)

    # Create the chart object
    chart = LineChart()
    chart.title = "Accès non autorisés au jeu"
    chart.style = 10
    chart.x_axis.title = "Année"
    chart.y_axis.title = "Accès non autorisés au jeu"

    # Loop through each row to add a series
    for row in range(2, ws.max_row + 1):
        values = Reference(ws, min_col=2, min_row=row,
                           max_row=row, max_col=ws.max_column)
        series = Series(values, title=ws.cell(row, 1).value)
        series.smooth = False  # Disable smoothed lines
        chart.series.append(series)

    # Define categories (X axis values) using the header row
    categories = Reference(ws, min_col=2, min_row=1,
                           max_row=1, max_col=ws.max_column)
    chart.set_categories(categories)

    # Set the size of the chart in points
    chart.width = 1300 / 72 * 2.54  # Width in cm
    chart.height = 800 / 72 * 2.54  # Height in cm

    # Ensure axis labels are displayed
    chart.y_axis.majorGridlines = ChartLines()
    chart.y_axis.minorGridlines = ChartLines()
    chart.y_axis.minorTickMark = "out"

    chart.x_axis.majorGridlines = ChartLines()
    chart.x_axis.minorGridlines = ChartLines()
    chart.x_axis.minorTickMark = "out"

    chart.x_axis.tickLblPos = 'low'
    chart.y_axis.tickLblPos = 'low'

    # Set number format for the axes to ensure values are displayed
    chart.x_axis.number_format = 'General'
    chart.y_axis.number_format = 'General'

    # Enable tick labels
    chart.x_axis.tickLblSkip = 1
    chart.y_axis.tickLblSkip = 1

    # Set the chart layout to include a title and legend
    chart.layout = Layout(
        manualLayout=ManualLayout(
            x=0.005,  # Position of the plot area from the left (0-1 scale)
            y=0.005,  # Position of the plot area from the top (0-1 scale)
            h=0.90,   # Height of the plot area (0-1 scale)
            w=0.85    # Width of the plot area (0-1 scale)
        )
    )

    # Position the chart on the new sheet
    ws_chart.add_chart(chart, "A1")

# The `with` statement ensures that the writer is closed and the file is saved
print(f"Workbook saved to {output_file_path}")
